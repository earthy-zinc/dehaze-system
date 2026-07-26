"""
文件生成器与解析器单元测试
"""
from __future__ import annotations

import csv
import io
from datetime import datetime

import pytest
from openpyxl import load_workbook

from app.core.code import ResultCode
from app.core.exceptions import BusinessException
from app.service.import_export.file_generator import write_csv, write_excel
from app.service.import_export.file_parser import (parse_csv, parse_excel,
                                                   validate_required_fields)
from app.service.import_export.models import ExportFieldConfig, ImportFieldConfig


def _make_export_fields() -> list[ExportFieldConfig]:
    return [
        ExportFieldConfig(field="username", label="用户名", order=2),
        ExportFieldConfig(field="nickname", label="昵称", order=1),
        ExportFieldConfig(field="secret", label="密钥", order=3, hidden=True),
    ]


class TestWriteExcel:
    def test_writes_headers_and_rows_sorted_by_order(self):
        fields = _make_export_fields()
        rows = [
            {"username": "u1", "nickname": "n1", "secret": "s1"},
            {"username": "u2", "nickname": "n2"},
        ]
        output = io.BytesIO()
        write_excel(fields, rows, output)

        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        # 表头按 order 排序：昵称, 用户名 (hidden 字段被排除)
        assert all_rows[0] == ("昵称", "用户名")
        assert all_rows[1] == ("n1", "u1")
        assert all_rows[2] == ("n2", "u2")

    def test_none_value_becomes_empty_string(self):
        fields = [ExportFieldConfig(field="username", label="用户名", order=1)]
        rows = [{"username": None}]
        output = io.BytesIO()
        write_excel(fields, rows, output)

        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        assert all_rows[1] == (None,)

    def test_date_format_applied(self):
        fields = [
            ExportFieldConfig(
                field="create_time",
                label="创建时间",
                order=1,
                date_format="%Y-%m-%d",
            )
        ]
        rows = [{"create_time": datetime(2026, 7, 27, 10, 30, 0)}]
        output = io.BytesIO()
        write_excel(fields, rows, output)

        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        assert all_rows[1] == ("2026-07-27",)


class TestWriteCsv:
    def test_writes_bom_and_headers_sorted_by_order(self):
        fields = _make_export_fields()
        rows = [{"username": "u1", "nickname": "n1"}]
        output = io.BytesIO()
        write_csv(fields, rows, output)

        content = output.getvalue()
        assert content.startswith("\ufeff".encode("utf-8"))
        text = content.decode("utf-8-sig")
        reader = list(csv.reader(io.StringIO(text)))
        # 按 order 排序：昵称, 用户名
        assert reader[0] == ["昵称", "用户名"]
        assert reader[1] == ["n1", "u1"]

    def test_none_value_becomes_empty_string(self):
        fields = [ExportFieldConfig(field="username", label="用户名", order=1)]
        rows = [{"username": None}]
        output = io.BytesIO()
        write_csv(fields, rows, output)

        text = output.getvalue().decode("utf-8-sig")
        reader = list(csv.reader(io.StringIO(text)))
        assert reader[1] == [""]


class TestParseExcel:
    def _make_excel_bytes(self, header: list[str], data_rows: list[list]) -> bytes:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(header)
        for row in data_rows:
            ws.append(row)
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def test_parse_maps_label_to_field(self):
        fields = [
            ImportFieldConfig(field="username", label="用户名", required=True),
            ImportFieldConfig(field="nickname", label="昵称"),
        ]
        content = self._make_excel_bytes(
            ["用户名", "昵称"], [["zhangsan", "张三"]]
        )
        rows = parse_excel(content, fields)

        assert len(rows) == 1
        assert rows[0]["username"] == "zhangsan"
        assert rows[0]["nickname"] == "张三"

    def test_parse_skips_empty_rows(self):
        fields = [ImportFieldConfig(field="username", label="用户名", required=True)]
        content = self._make_excel_bytes(
            ["用户名"], [["u1"], [None, None], ["u2"]]
        )
        rows = parse_excel(content, fields)
        assert len(rows) == 2

    def test_parse_missing_required_header_raises(self):
        fields = [
            ImportFieldConfig(field="username", label="用户名", required=True),
            ImportFieldConfig(field="nickname", label="昵称", required=True),
        ]
        content = self._make_excel_bytes(["用户名"], [["u1"]])
        with pytest.raises(BusinessException) as exc_info:
            parse_excel(content, fields)
        assert exc_info.value.code == ResultCode.IMPORT_TEMPLATE_MISMATCH

    def test_parse_ignores_unknown_columns(self):
        fields = [ImportFieldConfig(field="username", label="用户名", required=True)]
        content = self._make_excel_bytes(
            ["用户名", "未知列"], [["u1", "x"]]
        )
        rows = parse_excel(content, fields)
        assert rows[0] == {"username": "u1"}

    def test_parse_empty_file_raises(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        output = io.BytesIO()
        wb.save(output)
        fields = [ImportFieldConfig(field="username", label="用户名", required=True)]
        with pytest.raises(BusinessException) as exc_info:
            parse_excel(output.getvalue(), fields)
        assert exc_info.value.code == ResultCode.IMPORT_FILE_EMPTY


class TestParseCsv:
    def test_parse_maps_label_to_field(self):
        fields = [
            ImportFieldConfig(field="username", label="用户名", required=True),
            ImportFieldConfig(field="nickname", label="昵称"),
        ]
        content = "用户名,昵称\nzhangsan,张三\n".encode("utf-8")
        rows = parse_csv(content, fields)

        assert len(rows) == 1
        assert rows[0]["username"] == "zhangsan"
        assert rows[0]["nickname"] == "张三"

    def test_parse_handles_utf8_bom(self):
        fields = [ImportFieldConfig(field="username", label="用户名", required=True)]
        content = ("\ufeff用户名\nu1\n").encode("utf-8")
        rows = parse_csv(content, fields)
        assert rows[0]["username"] == "u1"

    def test_parse_handles_gbk_encoding(self):
        fields = [ImportFieldConfig(field="username", label="用户名", required=True)]
        content = "用户名\nu1\n".encode("gbk")
        rows = parse_csv(content, fields)
        assert rows[0]["username"] == "u1"

    def test_parse_missing_required_header_raises(self):
        fields = [
            ImportFieldConfig(field="username", label="用户名", required=True),
        ]
        content = "昵称\nn1\n".encode("utf-8")
        with pytest.raises(BusinessException) as exc_info:
            parse_csv(content, fields)
        assert exc_info.value.code == ResultCode.IMPORT_TEMPLATE_MISMATCH

    def test_parse_skips_blank_rows(self):
        fields = [ImportFieldConfig(field="username", label="用户名", required=True)]
        content = "用户名\nu1\n\nu2\n".encode("utf-8")
        rows = parse_csv(content, fields)
        assert len(rows) == 2


class TestValidateRequiredFields:
    def test_all_required_fields_present(self):
        fields = [
            ImportFieldConfig(field="username", label="用户名", required=True),
            ImportFieldConfig(field="nickname", label="昵称", required=True),
        ]
        row = {"username": "u1", "nickname": "n1"}
        assert validate_required_fields(row, fields) is None

    def test_missing_required_field(self):
        fields = [
            ImportFieldConfig(field="username", label="用户名", required=True),
        ]
        row = {"username": ""}
        result = validate_required_fields(row, fields)
        assert result is not None
        assert "用户名" in result

    def test_optional_field_skipped(self):
        fields = [
            ImportFieldConfig(field="username", label="用户名", required=True),
            ImportFieldConfig(field="email", label="邮箱", required=False),
        ]
        row = {"username": "u1", "email": ""}
        assert validate_required_fields(row, fields) is None
