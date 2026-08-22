import csv
import io
from contextlib import contextmanager
from datetime import datetime
from unittest.mock import AsyncMock, patch

from openpyxl import load_workbook

from app.service.import_export.handlers.user_export import UserExportHandler, _user_to_row
from app.service.import_export.handlers.user_import import UserImportHandler
from app.service.import_export.models import ExportContext, ImportOptions
from tests.stubs import NullDBSession

HASHED_PASSWORD = "$2b$12$abcdefghijklmnopqrstuvwxYz0123456789ABCDEFG"


def _callbacks(cancel=False):
    return AsyncMock(), AsyncMock(return_value=cancel)


@contextmanager
def _export_repos():
    with (
        patch("app.service.import_export.handlers.user_export.user_repository") as user_repo,
        patch("app.service.import_export.handlers.user_export.dept_repository") as dept_repo,
    ):
        user_repo.get_user_list = AsyncMock()
        dept_repo.get_children_ids = AsyncMock()
        yield user_repo, dept_repo


@contextmanager
def _import_repos(existing_usernames=None, create_side_effect=None):
    with (
        patch("app.service.import_export.handlers.user_import.user_repository") as user_repo,
        patch(
            "app.service.import_export.handlers.user_import.hash_password_async",
            new=AsyncMock(return_value=HASHED_PASSWORD),
        ),
    ):
        user_repo.get_existing_usernames = AsyncMock(return_value=existing_usernames or set())
        user_repo.create_user = AsyncMock(side_effect=create_side_effect)
        yield user_repo


class TestUserExportHandler:
    def test_get_module(self):
        assert UserExportHandler().get_module() == "user"

    def test_get_field_configs(self):
        handler = UserExportHandler()
        fields = handler.get_field_configs()
        assert len(fields) == 10
        assert [f.field for f in fields][:3] == ["id", "username", "nickname"]
        assert fields[3].field == "email"
        orders = [f.order for f in fields]
        assert orders == sorted(orders)
        create_time_field = next(f for f in fields if f.field == "create_time")
        assert create_time_field.date_format == "%Y-%m-%d %H:%M:%S"

    async def test_estimate_count_without_dept_id(self):
        handler = UserExportHandler()
        with _export_repos() as (user_repo, dept_repo):
            user_repo.get_user_list = AsyncMock(return_value=([], 42))
            count = await handler.estimate_count(NullDBSession(), {"keywords": "张"})
            assert count == 42
            dept_repo.get_children_ids.assert_not_called()
            user_repo.get_user_list.assert_awaited_once()
            call_kwargs = user_repo.get_user_list.call_args.kwargs
            assert call_kwargs["keywords"] == "张"
            assert call_kwargs["page"] == 1
            assert call_kwargs["page_size"] == 1

    async def test_estimate_count_with_dept_id(self):
        handler = UserExportHandler()
        with _export_repos() as (user_repo, dept_repo):
            dept_repo.get_children_ids = AsyncMock(return_value=[1, 2, 3])
            user_repo.get_user_list = AsyncMock(return_value=([], 100))
            count = await handler.estimate_count(NullDBSession(), {"deptId": 1})
            assert count == 100
            dept_repo.get_children_ids.assert_awaited_once()
            assert user_repo.get_user_list.call_args.kwargs["dept_ids"] == [1, 2, 3]

    async def test_export_writes_excel_with_rows(self):
        handler = UserExportHandler()
        users = [
            {
                "id": 1,
                "username": "=cmd|'/c calc'!A0",
                "nickname": "张\u200b三（研发）\r\n主程",
                "email": "",
                "mobile": "13800138000",
                "gender": 2,
                "status": 2,
                "deptName": "研发\u00a0部",
                "roleNames": "管理员,运维",
                "create_time": datetime(2026, 7, 27, 10, 30, 0),
            },
            {
                "id": 2,
                "username": "user０２",
                "nickname": "超长昵称" + "长" * 57,
                "email": None,
                "mobile": "",
                "gender": 1,
                "status": 1,
                "deptName": None,
                "roleNames": None,
                "create_time": None,
            },
        ]
        with _export_repos() as (user_repo, dept_repo):
            user_repo.get_user_list = AsyncMock(side_effect=[(users, 2), ([], 0)])
            output = io.BytesIO()
            ctx = ExportContext(
                task_id="t-20260727",
                module="user",
                format="excel",
                query_params={},
                total_count=2,
            )
            await handler.export(NullDBSession(), ctx, output, *_callbacks())
            output.seek(0)
            ws = load_workbook(output).active
            rows = list(ws.iter_rows(values_only=True))
            assert len(rows[0]) == 10
            assert rows[1][1] == "=cmd|'/c calc'!A0"
            assert rows[1][2] == "张\u200b三（研发）\r\n主程"
            assert rows[1][3] is None
            assert rows[1][5] == "女"
            assert rows[1][6] == "禁用"
            assert rows[1][7] == "研发\u00a0部"
            assert rows[1][9] == "2026-07-27 10:30:00"
            assert rows[2][1] == "user０２"
            assert rows[2][2] == "超长昵称" + "长" * 57
            assert rows[2][5] == "男"
            assert rows[2][6] == "正常"
            assert rows[2][7] is None
            assert rows[2][8] is None
            assert rows[2][9] is None

    async def test_export_csv_format(self):
        handler = UserExportHandler()
        users = [
            {
                "id": 1,
                "username": '=HYPERLINK("http://evil","点我")',
                "nickname": "张,三（全角，逗号）",
                "email": "   ",
                "mobile": "",
                "gender": 1,
                "status": 1,
                "deptName": "研发部",
                "roleNames": "管理员",
                "create_time": None,
            }
        ]
        with _export_repos() as (user_repo, dept_repo):
            user_repo.get_user_list = AsyncMock(side_effect=[(users, 1), ([], 0)])
            output = io.BytesIO()
            ctx = ExportContext(
                task_id="t1", module="user", format="csv", query_params={}, total_count=1
            )
            await handler.export(NullDBSession(), ctx, output, *_callbacks())
            content = output.getvalue()
            assert content.startswith("\ufeff".encode("utf-8"))
            reader = list(csv.reader(io.StringIO(content.decode("utf-8-sig"))))
            assert reader[0] == [
                "ID",
                "用户名",
                "昵称",
                "邮箱",
                "手机号",
                "性别",
                "状态",
                "部门",
                "角色",
                "创建时间",
            ]
            assert reader[1][1] == '=HYPERLINK("http://evil","点我")'
            assert reader[1][2] == "张,三（全角，逗号）"

    async def test_export_selected_fields_only(self):
        handler = UserExportHandler()
        users = [
            {
                "id": 1,
                "username": "zhangsan",
                "nickname": "张三",
                "email": "zhangsan@example.com",
            }
        ]
        with _export_repos() as (user_repo, dept_repo):
            user_repo.get_user_list = AsyncMock(side_effect=[(users, 1), ([], 0)])
            output = io.BytesIO()
            ctx = ExportContext(
                task_id="t1",
                module="user",
                format="excel",
                query_params={},
                total_count=1,
                selected_fields=["email", "username"],
            )
            await handler.export(NullDBSession(), ctx, output, *_callbacks())
            output.seek(0)
            ws = load_workbook(output).active
            rows = list(ws.iter_rows(values_only=True))
            assert rows[0] == ("用户名", "邮箱")
            assert rows[1] == ("zhangsan", "zhangsan@example.com")

    async def test_export_stops_on_cancel(self):
        handler = UserExportHandler()
        users = [
            {
                "id": 1,
                "username": "zhangsan",
                "nickname": "张三",
                "gender": 1,
                "status": 1,
            }
        ]
        with _export_repos() as (user_repo, dept_repo):
            user_repo.get_user_list = AsyncMock(return_value=(users, 1))
            progress_cb, cancel_cb = _callbacks(cancel=True)
            output = io.BytesIO()
            ctx = ExportContext(
                task_id="t1", module="user", format="excel", query_params={}, total_count=1
            )
            await handler.export(NullDBSession(), ctx, output, progress_cb, cancel_cb)
            assert user_repo.get_user_list.await_count == 1
            cancel_cb.assert_awaited_once()


class TestUserImportHandler:
    def test_get_module(self):
        assert UserImportHandler().get_module() == "user"

    def test_get_field_configs(self):
        handler = UserImportHandler()
        fields = handler.get_field_configs()
        assert len(fields) == 8
        username_field = next(f for f in fields if f.field == "username")
        assert username_field.required is True
        assert username_field.max_length == 64
        nickname_field = next(f for f in fields if f.field == "nickname")
        assert nickname_field.required is True

    def test_get_template_sample_data(self):
        handler = UserImportHandler()
        samples = handler.get_template_sample_data()
        assert len(samples) == 1
        sample = samples[0]
        assert sample["username"] == "zhangsan"
        assert sample["nickname"] == "张三"
        assert sample["mobile"] == "13800138000"

    async def test_import_batch_success(self):
        handler = UserImportHandler()
        rows = [
            {
                "username": "  zhangsan  ",
                "nickname": "张三（研发）",
                "gender": "男",
                "mobile": "13800138000",
                "email": "zhangsan@example.com",
            }
        ]
        with _import_repos() as user_repo:
            result = await handler.import_batch(
                NullDBSession(),
                rows,
                ImportOptions(mode="all"),
                *_callbacks(),
            )
            assert result.total_rows == 1
            assert result.success_count == 1
            assert result.failure_count == 0
            user_repo.create_user.assert_awaited_once()
            created_user = user_repo.create_user.call_args.args[1]
            assert created_user.username == "zhangsan"
            assert created_user.nickname == "张三（研发）"
            assert created_user.password == HASHED_PASSWORD

    async def test_import_batch_duplicate_username_skipped(self):
        handler = UserImportHandler()
        rows = [
            {"username": "zhangsan", "nickname": "张三", "mobile": "13800138000"},
            {"username": "lisi", "nickname": "李四", "mobile": "13900139000"},
        ]
        with _import_repos(existing_usernames={"zhangsan"}) as user_repo:
            result = await handler.import_batch(
                NullDBSession(),
                rows,
                ImportOptions(mode="partial"),
                *_callbacks(),
            )
            assert result.total_rows == 2
            assert result.success_count == 1
            assert result.failure_count == 1
            assert len(result.errors) == 1
            assert result.errors[0].row == 2
            user_repo.create_user.assert_awaited_once()

    async def test_import_batch_blank_username_recorded_as_error(self):
        handler = UserImportHandler()
        rows = [{"username": "   ", "nickname": "张三"}]
        with _import_repos() as user_repo:
            result = await handler.import_batch(
                NullDBSession(),
                rows,
                ImportOptions(mode="partial"),
                *_callbacks(),
            )
            assert result.failure_count == 1
            assert "为空" in result.errors[0].message
            user_repo.create_user.assert_not_called()

    async def test_import_batch_duplicate_in_batch_skipped(self):
        handler = UserImportHandler()
        rows = [
            {"username": "zhangsan", "nickname": "张三", "mobile": "13800138000"},
            {"username": "zhangsan", "nickname": "张三同学", "mobile": "13900139000"},
        ]
        with _import_repos() as user_repo:
            result = await handler.import_batch(
                NullDBSession(),
                rows,
                ImportOptions(mode="partial"),
                *_callbacks(),
            )
            assert result.success_count == 1
            assert result.failure_count == 1
            assert result.errors[0].row == 3

    async def test_import_batch_default_dept_id_from_options(self):
        handler = UserImportHandler()
        rows = [{"username": "zhangsan", "nickname": "张三", "mobile": "13800138000"}]
        with _import_repos() as user_repo:
            await handler.import_batch(
                NullDBSession(),
                rows,
                ImportOptions(mode="all", extra={"deptId": 100}),
                *_callbacks(),
            )
            created_user = user_repo.create_user.call_args.args[1]
            assert created_user.dept_id == 100

    async def test_import_batch_create_exception_recorded(self):
        handler = UserImportHandler()
        rows = [{"username": "zhangsan", "nickname": "张三", "mobile": "13800138000"}]
        with _import_repos(create_side_effect=RuntimeError("DB 错误")) as user_repo:
            result = await handler.import_batch(
                NullDBSession(),
                rows,
                ImportOptions(mode="partial"),
                *_callbacks(),
            )
            assert result.failure_count == 1
            assert "DB 错误" in result.errors[0].message

    async def test_import_batch_gender_female(self):
        handler = UserImportHandler()
        rows = [{"username": "zhangsan", "nickname": "张三", "gender": "女", "mobile": "13800138000"}]
        with _import_repos() as user_repo:
            await handler.import_batch(
                NullDBSession(),
                rows,
                ImportOptions(mode="all"),
                *_callbacks(),
            )
            created_user = user_repo.create_user.call_args.args[1]
            assert created_user.gender == 2

    async def test_import_batch_invalid_gender_rejected(self):
        handler = UserImportHandler()
        rows = [{"username": "zhangsan", "nickname": "张三", "gender": "未知"}]
        with _import_repos() as user_repo:
            result = await handler.import_batch(
                NullDBSession(),
                rows,
                ImportOptions(mode="all"),
                *_callbacks(),
            )
            assert result.success_count == 0
            assert result.failure_count == 1
            assert "性别取值无效" in result.errors[0].message
            user_repo.create_user.assert_not_called()

    async def test_import_batch_role_ids_parsed(self):
        handler = UserImportHandler()
        rows = [
            {
                "username": "zhangsan",
                "nickname": "张三",
                "mobile": "13800138000",
                "role_ids": "1, 2, 3",
            }
        ]
        with _import_repos() as user_repo:
            await handler.import_batch(
                NullDBSession(),
                rows,
                ImportOptions(mode="all"),
                *_callbacks(),
            )
            role_ids = user_repo.create_user.call_args.args[2]
            assert role_ids == [1, 2, 3]

    async def test_import_batch_fullwidth_role_ids_recorded_as_error(self):
        handler = UserImportHandler()
        rows = [
            {
                "username": "zhangsan",
                "nickname": "张三",
                "mobile": "13800138000",
                "role_ids": "1，2,3",
            }
        ]
        with _import_repos() as user_repo:
            result = await handler.import_batch(
                NullDBSession(),
                rows,
                ImportOptions(mode="partial"),
                *_callbacks(),
            )
            assert result.failure_count == 1
            assert result.success_count == 0
            assert result.errors[0].row == 2
            user_repo.create_user.assert_not_called()


class TestUserToRow:
    def test_gender_follows_dict_convention(self):
        assert _user_to_row({"gender": 1, "status": 1})["gender_label"] == "男"
        assert _user_to_row({"gender": 2, "status": 1})["gender_label"] == "女"
        assert _user_to_row({"gender": 0, "status": 1})["gender_label"] == "未知"
        assert _user_to_row({"status": 1})["gender_label"] == "未知"

    def test_disabled_status_not_defaulted_to_normal(self):
        assert _user_to_row({"gender": 1, "status": 0})["status_label"] == "禁用"
        assert _user_to_row({"gender": 1, "status": 2})["status_label"] == "禁用"
        assert _user_to_row({"gender": 1, "status": 1})["status_label"] == "正常"
