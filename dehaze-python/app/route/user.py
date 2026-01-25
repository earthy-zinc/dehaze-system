"""
用户管理路由 - 使用 flask-openapi3 自动生成 Swagger 文档
"""
from flask_openapi3 import APIBlueprint, Tag
from werkzeug.datastructures import FileStorage

from app.models.schema.user import (
    UserPageQuery,
    UserStatusQuery,
    UserIdPath,
    UserIdsPath,
    LoginForm,
    RegisterForm,
    UserForm,
    PasswordForm,
)
from app.service.user import UserService
from app.utils.jwt_util import jwt_required, get_current_user_id
from app.utils.result import success, error


# 定义 Tag
user_tag = Tag(name="用户管理", description="用户相关接口")

# 创建 APIBlueprint（自动携带安全配置）
user_blueprint = APIBlueprint(
    "user",
    __name__,
    url_prefix="/api/v1/users",
    abp_tags=[user_tag],
    abp_security=[{"BearerAuth": []}]
)


@user_blueprint.post(
    "/login",
    summary="用户登录",
    description="用户登录接口",
    security=None
)
def login(body: LoginForm):
    """用户登录"""
    user = UserService.authenticate_user(body.username, body.password)
    if not user:
        return error('用户名或密码错误', 401)

    if user.status != 1:
        return error('用户已被禁用', 401)

    token = UserService.generate_token(user.id)

    return success({
        'token': token,
        'user': {
            'id': user.id,
            'username': user.username,
            'nickname': user.nickname
        }
    })


@user_blueprint.post(
    "/register",
    summary="用户注册",
    description="用户注册接口",
    security=None
)
def register(body: RegisterForm):
    """用户注册"""
    # 检查用户名是否已存在
    existing_user = UserService.get_user_by_username(body.username)
    if existing_user:
        return error('用户名已存在', 400)

    # 创建用户
    user = UserService.create_user(body.username, body.password, body.nickname)

    return success({
        'id': user.id,
        'username': user.username,
        'nickname': user.nickname
    }, '注册成功')


@user_blueprint.get(
    "/me",
    summary="获取当前用户信息",
    description="获取当前登录用户的信息"
)
@jwt_required
def get_current_user():
    """获取当前用户信息"""
    user_id = get_current_user_id()
    user = UserService.get_user_by_id(user_id)

    if not user:
        return error('用户不存在', 404)

    # 获取用户角色
    roles = UserService.get_user_roles(user_id)
    role_list = [role.code for role in roles]

    # 获取用户权限
    perms = UserService.get_user_permissions(user_id)

    return success({
        'userId': user.id,
        'username': user.username,
        'nickname': user.nickname,
        'avatar': user.avatar,
        'roles': role_list,
        'perms': perms
    })


@user_blueprint.get(
    "/page",
    summary="获取用户分页列表",
    description="获取用户分页列表"
)
@jwt_required
def get_user_page(query: UserPageQuery):
    """获取用户分页列表"""
    users, total = UserService.get_user_list(
        query.pageNum,
        query.pageSize,
        None,
        query.keywords,
        query.status,
        query.deptId,
        query.startTime,
        query.endTime
    )

    user_list = []
    for user in users:
        # 添加 genderLabel 和 statusLabel 字段
        gender_label = '男' if user.gender == 1 else '女'
        status_label = '正常' if user.status == 1 else '禁用'

        user_list.append({
            'id': user.id,
            'username': user.username,
            'nickname': user.nickname,
            'mobile': user.mobile,
            'genderLabel': gender_label,
            'avatar': user.avatar,
            'status': user.status,
            'statusLabel': status_label,
            'email': user.email,
            'createTime': user.create_time.strftime('%Y-%m-%d %H:%M:%S') if user.create_time else None
        })

    return success({
        'list': user_list,
        'total': total,
        'pageNum': query.pageNum,
        'pageSize': query.pageSize
    })


@user_blueprint.post(
    "/",
    summary="新增用户",
    description="新增用户接口"
)
@jwt_required
def create_user(body: UserForm):
    """新增用户"""
    data = body.model_dump(exclude_none=True)
    result = UserService.create_user_with_roles(data)

    if result.get('error'):
        return error(result['error'], 400)

    return success(result['data'], '新增成功')


@user_blueprint.get(
    "/<int:user_id>/form",
    summary="获取用户表单数据",
    description="获取用户表单数据"
)
@jwt_required
def get_user_form(path: UserIdPath):
    """获取用户表单数据"""
    # 无论用户是否存在，都返回空数据，与Java后端保持一致
    user_data = UserService.get_user_form_data(path.user_id)

    # 如果用户存在，添加 avatar 字段
    if user_data:
        user = UserService.get_user_by_id(path.user_id)
        if user:
            user_data['avatar'] = user.avatar

    return success(user_data)


@user_blueprint.put(
    "/<int:user_id>",
    summary="更新用户",
    description="更新用户信息"
)
@jwt_required
def update_user(path: UserIdPath, body: UserForm):
    """更新用户信息"""
    data = body.model_dump(exclude_none=True)
    result = UserService.update_user_with_roles(path.user_id, data)

    if result.get('error'):
        return error(result['error'], 400)

    return success(result.get('data'), '更新成功')


@user_blueprint.patch(
    "/<int:user_id>/status",
    summary="更新用户状态",
    description="更新用户状态（启用/禁用）"
)
@jwt_required
def update_user_status(path: UserIdPath, query: UserStatusQuery):
    """更新用户状态"""
    result = UserService.update_user_status(path.user_id, query.status)

    if not result:
        return error('用户不存在', 404)

    return success(None, '更新成功')


@user_blueprint.put(
    "/<int:user_id>/password",
    summary="修改用户密码",
    description="修改用户密码"
)
@jwt_required
def update_password(path: UserIdPath, body: PasswordForm):
    """修改用户密码"""
    result = UserService.update_password(path.user_id, body.password)

    if not result:
        return error('用户不存在', 404)

    return success(None, '修改成功')


@user_blueprint.delete(
    "/<ids>",
    summary="删除用户",
    description="删除用户（逻辑删除，支持批量删除）"
)
@jwt_required
def delete_user(path: UserIdsPath):
    """删除用户（支持批量删除）"""
    # 支持批量删除，多个ID用逗号分隔
    id_list = [int(id_str.strip()) for id_str in path.ids.split(',') if id_str.strip()]

    result_count = 0
    for user_id in id_list:
        result = UserService.delete_user(user_id)
        if result:
            result_count += 1

    return success(None, f'成功删除{result_count}个用户')


@user_blueprint.get(
    "/template",
    summary="用户导入模板下载",
    description="下载用户导入模板"
)
@jwt_required
def download_template():
    """下载用户导入模板"""
    from flask import make_response
    import openpyxl
    from io import BytesIO

    # 创建一个新的Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "用户导入模板"

    # 设置表头
    headers = ['用户名', '昵称', '密码', '邮箱', '手机号', '性别', '部门ID', '角色ID(多个用逗号分隔)']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 返回文件
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=user_import_template.xlsx'

    return response


@user_blueprint.get(
    "/export",
    summary="导出用户",
    description="导出用户数据"
)
@jwt_required
def export_users(query: UserPageQuery):
    """导出用户数据"""
    from flask import make_response
    import openpyxl
    from io import BytesIO

    # 获取用户列表（不分页，导出所有数据）
    users, total = UserService.get_user_list(
        page=1,
        page_size=10000,
        username=None,
        keywords=query.keywords,
        status=query.status,
        dept_id=query.deptId,
        create_time_start=query.startTime,
        create_time_end=query.endTime
    )

    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "用户列表"

    # 设置表头
    headers = ['ID', '用户名', '昵称', '邮箱', '手机号', '性别', '状态', '部门ID']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 填充数据
    for row, user in enumerate(users, 2):
        # 处理可能为None或字节类型的字段
        gender_value = int(user.gender) if user.gender is not None else 1
        status_value = int(user.status) if user.status is not None else 1
        gender_label = '男' if gender_value == 1 else '女'
        status_label = '正常' if status_value == 1 else '禁用'

        ws.cell(row=row, column=1, value=user.id)
        ws.cell(row=row, column=2, value=user.username)
        ws.cell(row=row, column=3, value=user.nickname)
        ws.cell(row=row, column=4, value=user.email or '')
        ws.cell(row=row, column=5, value=user.mobile or '')
        ws.cell(row=row, column=6, value=gender_label)
        ws.cell(row=row, column=7, value=status_label)
        ws.cell(row=row, column=8, value=user.dept_id or '')

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 返回文件
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=users_export.xlsx'

    return response


@user_blueprint.post(
    "/import",
    summary="导入用户",
    description="通过Excel批量导入用户"
)
@jwt_required
def import_users(file: FileStorage = None, deptId: int = None):
    """导入用户"""
    import openpyxl

    if not file:
        return error('请选择要导入的文件', 400)

    if not deptId:
        return error('请选择目标部门', 400)

    # 检查文件格式
    if not file.filename.endswith(('.xls', '.xlsx')):
        return error('仅支持.xls和.xlsx格式的文件', 400)

    try:
        # 读取Excel文件
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # 执行导入
        result = UserService.import_users(ws, deptId)

        return success(result, f'导入完成，成功{result["successCount"]}条，失败{result["failedCount"]}条')
    except Exception as e:
        return error(f'导入失败: {str(e)}', 500)
