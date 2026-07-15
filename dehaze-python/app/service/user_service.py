"""
用户服务

提供用户 CRUD 功能，支持角色分配、Excel 导入导出等
"""

import re
import secrets
import string
from io import BytesIO
from typing import Any

import openpyxl
from app.config import settings
from app.core.exceptions import BusinessException
from app.models.entity.sys_user import SysUser
from app.repository.dept_repository import dept_repository
from app.repository.user_repository import user_repository
from app.utils.password import hash_password_async
from sqlalchemy.ext.asyncio import AsyncSession


def validate_password_complexity(password: str) -> tuple[bool, str]:
    """
    验证密码复杂度

    设计文档要求：最小长度 8 位，复杂度为中（至少包含字母和数字）

    Args:
        password: 待验证的密码

    Returns:
        (是否通过, 错误信息)
    """
    if len(password) < settings.PASSWORD_MIN_LENGTH:
        return False, f"密码长度不能少于 {settings.PASSWORD_MIN_LENGTH} 位"

    if settings.PASSWORD_REQUIRE_COMPLEXITY:
        has_letter = bool(re.search(r"[a-zA-Z]", password))
        has_digit = bool(re.search(r"\d", password))

        if not (has_letter and has_digit):
            return False, "密码必须包含字母和数字"

    return True, ""


def generate_random_password(length: int = 12) -> str:
    """生成随机密码，确保包含字母和数字"""
    if length < 4:
        length = 4

    # 确保至少包含一个字母和一个数字
    password_chars = [
        secrets.choice(string.ascii_lowercase),
        secrets.choice(string.ascii_uppercase),
        secrets.choice(string.digits),
        secrets.choice("!@#$%^&*"),
    ]

    # 剩余字符从完整字符集中随机选择
    alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
    remaining_length = length - len(password_chars)
    password_chars.extend(secrets.choice(alphabet)
                          for _ in range(remaining_length))

    # 打乱字符顺序（使用密码学安全的 SystemRandom）
    secrets.SystemRandom().shuffle(password_chars)

    return "".join(password_chars)


class UserService:
    """用户服务（异步版本）"""

    @staticmethod
    async def get_user_list(
        db: AsyncSession,
        page: int,
        page_size: int,
        keywords: str | None = None,
        status: int | None = None,
        dept_id: int | None = None,
        create_time_start: str | None = None,
        create_time_end: str | None = None,
    ) -> tuple[list[SysUser], int]:
        """
        获取用户列表（分页）

        Args:
            db: 异步数据库会话
            page: 页码
            page_size: 每页数量
            keywords: 关键词搜索
            status: 状态筛选
            dept_id: 部门ID筛选（包含下级部门）
            create_time_start: 创建时间开始
            create_time_end: 创建时间结束

        Returns:
            (用户列表, 总数)
        """
        dept_ids = None
        if dept_id:
            dept_ids = await dept_repository.get_children_ids(db, dept_id)

        return await user_repository.get_user_list(
            db,
            page=page,
            page_size=page_size,
            keywords=keywords,
            status=status,
            dept_ids=dept_ids,
            create_time_start=create_time_start,
            create_time_end=create_time_end,
        )

    @staticmethod
    async def get_user_form_data(db: AsyncSession, user_id: int) -> dict[str, Any] | None:
        """
        获取用户表单数据

        Args:
            db: 异步数据库会话
            user_id: 用户ID

        Returns:
            用户表单数据
        """
        user = await user_repository.get_by_id(db, user_id)
        if not user:
            return None

        role_ids = await user_repository.get_user_role_ids(db, user_id)

        return {
            "id": user.id,
            "username": user.username,
            "nickname": user.nickname,
            "gender": user.gender,
            "deptId": user.dept_id,
            "mobile": user.mobile,
            "email": user.email,
            "status": user.status,
            "avatar": user.avatar,
            "roleIds": role_ids,
        }

    @staticmethod
    async def create_user_with_roles(
        db: AsyncSession,
        data: dict[str, Any],
    ) -> SysUser:
        """
        创建新用户并关联角色

        Args:
            db: 异步数据库会话
            data: 用户数据

        Returns:
            创建的用户对象

        Raises:
            BusinessException: 用户名为空或用户名已存在
        """
        username = data.get("username")
        nickname = data.get("nickname", username)
        gender = data.get("gender")
        dept_id = data.get("deptId")
        mobile = data.get("mobile")
        email = data.get("email")
        status = data.get("status", 1)
        role_ids = data.get("roleIds", [])

        if not username:
            raise BusinessException("用户名不能为空")

        existing_user = await user_repository.get_by_username(db, username)
        if existing_user:
            raise BusinessException("用户名已存在")

        # 使用配置的默认密码
        plain_password = settings.DEFAULT_PASSWORD
        hashed_password = await hash_password_async(plain_password)

        user = SysUser(
            username=username,
            nickname=nickname,
            gender=gender,
            dept_id=dept_id,
            mobile=mobile,
            email=email,
            password=hashed_password,
            status=status,
        )

        user = await user_repository.create_user(db, user, role_ids)
        await db.commit()
        return user

    @staticmethod
    async def update_user_with_roles(
        db: AsyncSession,
        user_id: int,
        data: dict[str, Any],
    ) -> None:
        """
        更新用户信息并关联角色

        Args:
            db: 异步数据库会话
            user_id: 用户ID
            data: 用户数据

        Raises:
            BusinessException: 用户不存在或用户名已存在
        """
        user = await user_repository.get_by_id(db, user_id)
        if not user:
            raise BusinessException("用户不存在")

        username = data.get("username")
        nickname = data.get("nickname")
        gender = data.get("gender")
        dept_id = data.get("deptId")
        mobile = data.get("mobile")
        email = data.get("email")
        role_ids = data.get("roleIds", [])
        status = data.get("status")

        # 用户名冲突校验（排除当前用户）
        if username is not None and username != user.username:
            exists = await user_repository.check_username_exists(
                db, username, exclude_id=user_id
            )
            if exists:
                raise BusinessException("用户名已存在")
            user.username = username

        if nickname is not None:
            user.nickname = nickname
        if gender is not None:
            user.gender = gender
        if dept_id is not None:
            user.dept_id = dept_id
        if mobile is not None:
            user.mobile = mobile
        if email is not None:
            user.email = email
        if status is not None:
            user.status = status

        await db.flush()
        await user_repository.replace_user_roles(db, user_id, role_ids)
        await db.commit()

    @staticmethod
    async def update_user_status(
        db: AsyncSession,
        user_id: int,
        status: int,
    ) -> None:
        """
        更新用户状态

        Args:
            db: 异步数据库会话
            user_id: 用户ID
            status: 状态（1-正常，0-禁用）

        Raises:
            BusinessException: 用户不存在或为超级管理员
        """
        user = await user_repository.get_by_id(db, user_id)
        if not user:
            raise BusinessException("用户不存在")

        if user.username == "root":
            raise BusinessException("超级管理员不可禁用")

        user.status = status
        await db.commit()

    @staticmethod
    async def update_password(
        db: AsyncSession,
        user_id: int,
        new_password: str,
    ) -> None:
        """
        更新用户密码

        Args:
            db: 异步数据库会话
            user_id: 用户ID
            new_password: 新密码

        Raises:
            BusinessException: 用户不存在或密码复杂度不符合要求
        """
        # 验证密码复杂度
        is_valid, error_msg = validate_password_complexity(new_password)
        if not is_valid:
            raise BusinessException(error_msg)

        user = await user_repository.get_by_id(db, user_id)
        if not user:
            raise BusinessException("用户不存在")

        hashed_password = await hash_password_async(new_password)
        user.password = hashed_password
        await db.commit()

    @staticmethod
    async def delete_users(db: AsyncSession, ids: str) -> dict[str, int]:
        """
        删除用户（逻辑删除，支持批量）

        Args:
            db: 异步数据库会话
            ids: 用户ID，多个以英文逗号分隔

        Returns:
            删除统计 {"deleted_count": int, "protected_count": int}

        Raises:
            BusinessException: 未指定要删除的用户
        """
        user_ids = [int(id_str.strip())
                    for id_str in ids.split(",") if id_str.strip()]

        if not user_ids:
            raise BusinessException("未指定要删除的用户")

        protected_ids = await user_repository.get_protected_user_ids(db, user_ids)

        ids_to_delete = [uid for uid in user_ids if uid not in protected_ids]

        if ids_to_delete:
            await user_repository.soft_delete_by_ids(db, ids_to_delete)
            await db.commit()

        return {"deleted_count": len(ids_to_delete), "protected_count": len(protected_ids)}

    @staticmethod
    async def import_users(
        db: AsyncSession,
        worksheet,
        dept_id: int | None = None,
    ) -> dict[str, Any]:
        """
        从 Excel 工作表导入用户

        Args:
            db: 异步数据库会话
            worksheet: openpyxl 工作表对象
            dept_id: 目标部门ID

        Returns:
            导入结果，包含成功数、失败数和失败明细
        """
        success_count = 0
        failed_count = 0
        failures = []

        # 读取所有数据行
        rows = list(worksheet.iter_rows(min_row=2, values_only=True))

        # 第一遍：收集所有用户名，批量查询已存在的（避免 N+1）
        all_usernames = [
            str(row[0]) for row in rows if row and row[0]
        ]
        existing_usernames = await user_repository.get_existing_usernames(
            db, all_usernames
        )

        # 第二遍：逐行创建用户
        seen_usernames: set[str] = set()
        for row_num, row in enumerate(rows, start=2):
            try:
                username = row[0]
                nickname = row[1]
                password = row[2]
                email = row[3]
                mobile = row[4]
                gender = row[5]
                user_dept_id = row[6] if row[6] else dept_id
                role_ids_str = row[7]

                if not username or not nickname:
                    failures.append({
                        "row": row_num,
                        "username": username or "",
                        "reason": "用户名或昵称不能为空",
                    })
                    failed_count += 1
                    continue

                username_str = str(username)

                # 检查数据库已存在
                if username_str in existing_usernames:
                    failures.append({
                        "row": row_num,
                        "username": username_str,
                        "reason": "用户名已存在",
                    })
                    failed_count += 1
                    continue

                # 检查本次导入批次内重复
                if username_str in seen_usernames:
                    failures.append({
                        "row": row_num,
                        "username": username_str,
                        "reason": "用户名在导入文件中重复",
                    })
                    failed_count += 1
                    continue

                seen_usernames.add(username_str)

                if gender == "男":
                    gender_value = 1
                elif gender == "女":
                    gender_value = 0
                else:
                    gender_value = 1

                role_ids = []
                if role_ids_str:
                    role_ids = [
                        int(rid.strip())
                        for rid in str(role_ids_str).split(",")
                        if rid.strip()
                    ]

                # 使用提供的密码或默认密码
                plain_password = str(
                    password) if password else settings.DEFAULT_PASSWORD
                hashed_password = await hash_password_async(plain_password)

                user = SysUser(
                    username=username_str,
                    nickname=str(nickname),
                    password=hashed_password,
                    email=str(email) if email else None,
                    mobile=str(mobile) if mobile else None,
                    gender=gender_value,
                    dept_id=int(user_dept_id) if user_dept_id else None,
                    status=1,
                )

                await user_repository.create_user(db, user, role_ids)
                success_count += 1

            except Exception as e:
                failures.append({
                    "row": row_num,
                    "username": str(row[0]) if row else "",
                    "reason": str(e),
                })
                failed_count += 1

        await db.commit()
        return {
            "successCount": success_count,
            "failedCount": failed_count,
            "failures": failures,
        }

    @staticmethod
    def generate_import_template() -> BytesIO:
        """
        生成用户导入模板

        Returns:
            Excel 文件字节流
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("用户导入模板")
        else:
            ws.title = "用户导入模板"

        headers = ["用户名", "昵称", "密码", "邮箱",
                   "手机号", "性别", "部门ID", "角色ID(多个用逗号分隔)"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    async def export_users(
        db: AsyncSession,
        keywords: str | None = None,
        status: int | None = None,
        dept_id: int | None = None,
        create_time_start: str | None = None,
        create_time_end: str | None = None,
    ) -> BytesIO:
        """
        导出用户数据到 Excel

        Args:
            db: 异步数据库会话
            keywords: 关键词搜索
            status: 状态筛选
            dept_id: 部门ID筛选
            create_time_start: 创建时间开始
            create_time_end: 创建时间结束

        Returns:
            Excel 文件字节流
        """
        dept_ids = None
        if dept_id:
            dept_ids = await dept_repository.get_children_ids(db, dept_id)

        users, _ = await user_repository.get_user_list(
            db,
            page=1,
            page_size=10000,
            keywords=keywords,
            status=status,
            dept_ids=dept_ids,
            create_time_start=create_time_start,
            create_time_end=create_time_end,
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("用户列表")
        else:
            ws.title = "用户列表"

        headers = ["ID", "用户名", "昵称", "邮箱", "手机号", "性别", "状态", "部门ID"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        for row, user in enumerate(users, 2):
            gender_value = int(user["gender"]) if user.get("gender") is not None else 1
            status_value = int(user["status"]) if user.get("status") is not None else 1
            gender_label = "男" if gender_value == 1 else "女"
            status_label = "正常" if status_value == 1 else "禁用"

            ws.cell(row=row, column=1, value=user["id"])
            ws.cell(row=row, column=2, value=user["username"])
            ws.cell(row=row, column=3, value=user["nickname"])
            ws.cell(row=row, column=4, value=user.get("email") or "")
            ws.cell(row=row, column=5, value=user.get("mobile") or "")
            ws.cell(row=row, column=6, value=gender_label)
            ws.cell(row=row, column=7, value=status_label)
            ws.cell(row=row, column=8, value=user.get("dept_id") or "")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output


# 单例
user_service = UserService()
