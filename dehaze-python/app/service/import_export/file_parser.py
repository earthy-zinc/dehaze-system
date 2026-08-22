"""
导入文件解析器（Excel/CSV）
"""

from __future__ import annotations

import csv
import io

from openpyxl import load_workbook

from app.core.code import ResultCode
from app.core.exceptions import BusinessException
from app.service.import_export.models import ImportFieldConfig


def parse_excel(
    content: bytes,
    fields: list[ImportFieldConfig],
) -> list[dict]:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise BusinessException(
            ResultCode.IMPORT_FILE_PARSE_ERROR, f"Excel 解析失败: {e}"
        ) from None
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise BusinessException(ResultCode.IMPORT_FILE_EMPTY)
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    _validate_header(header, fields)
    return _map_rows(header, fields, rows, missing_cell=None)


def parse_csv(
    content: bytes,
    fields: list[ImportFieldConfig],
) -> list[dict]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = content.decode("gbk")
        except UnicodeDecodeError as e:
            raise BusinessException(
                ResultCode.IMPORT_FILE_PARSE_ERROR, f"CSV 编码解析失败: {e}"
            ) from None
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise BusinessException(ResultCode.IMPORT_FILE_EMPTY)
    header = [c.strip() for c in rows[0]]
    _validate_header(header, fields)
    return _map_rows(header, fields, rows, missing_cell="")


def _map_rows(
    header: list[str],
    fields: list[ImportFieldConfig],
    rows: list[list],
    missing_cell,
) -> list[dict]:
    label_to_field = {f.label: f for f in fields}
    data_rows: list[dict] = []
    for row in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        row_dict: dict = {}
        for idx, label in enumerate(header):
            field_cfg = label_to_field.get(label)
            if field_cfg is None:
                continue
            row_dict[field_cfg.field] = row[idx] if idx < len(row) else missing_cell
        data_rows.append(row_dict)
    return data_rows


def _validate_header(header: list[str], fields: list[ImportFieldConfig]) -> None:
    actual = set(header)
    missing = [f for f in fields if f.required and f.label not in actual]
    if missing:
        raise BusinessException(
            ResultCode.IMPORT_TEMPLATE_MISMATCH,
            f"模板缺少必填列: {','.join(f.label for f in missing)}",
        )


def validate_required_fields(row: dict, fields: list[ImportFieldConfig]) -> str | None:
    for f in fields:
        if f.required:
            v = row.get(f.field)
            if v is None or (isinstance(v, str) and not v.strip()):
                return f"必填字段 [{f.label}] 为空"
    return None
